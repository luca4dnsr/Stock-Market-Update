"""
excel_writer.py — 원본 Excel 레이아웃 재현
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────
# 스타일 상수
# ──────────────────────────────────────────────────────────
POS_FONT      = Font(color="CC0000", bold=True)         # 빨강 (상승)
NEG_FONT      = Font(color="0000CC", bold=True)         # 파랑 (하락)
FLAT_FONT     = Font(color="444444")
HEADER_FONT   = Font(bold=True, size=10)
TITLE_FONT    = Font(bold=True, size=22, color="222222")
DATE_FONT     = Font(bold=True, size=10, color="555555")
ADV_FONT      = Font(bold=True, color="CC0000", size=11)
DEC_FONT      = Font(bold=True, color="0000CC", size=11)

HEADER_FILL   = PatternFill("solid", fgColor="D0D0D0")
TOP_ROW_FILL  = PatternFill("solid", fgColor="FFF5E6")   # 연한 주황
BOT_ROW_FILL  = PatternFill("solid", fgColor="E6EEFF")   # 연한 파랑
SPX_FILL      = PatternFill("solid", fgColor="EEEEEE")
SECT_BG       = PatternFill("solid", fgColor="F5F5F5")

THIN_SIDE   = Side(style="thin",   color="BBBBBB")
THICK_SIDE  = Side(style="medium", color="888888")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE, bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

# 섹터 2열 배치 (좌5 + 우6)
SECTOR_LEFT_COLS  = [
    "Financials",
    "Information Technology",
    "Materials",
    "Consumer Discretionary",
    "Health Care",
]
SECTOR_RIGHT_COLS = [
    "Industrials",
    "Energy",
    "Utilities",
    "Communication Services",
    "Consumer Staples",
    "Real Estate",
]

# 열 너비 (문자 단위)
COL_WIDTHS = {
    "A": 6,   # rank
    "B": 12,  # ticker
    "C": 28,  # name
    "D": 12,  # market cap
    "E": 8,   # 1d
    "F": 8,   # 1w
    "G": 9,   # 1m
    "H": 9,   # 3m
    "I": 9,   # mc_rank
    "J": 26,  # sector
    "K": 46,  # business summary
    "L": 62,  # move reason
}


# ──────────────────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────────────────

def _pct_str(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "—"
    return f"{val:+.1f}"


def _set_pct_cell(cell, val):
    """수익률 셀에 값, 색상, 정렬 적용."""
    cell.value = _pct_str(val)
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    if val is None or (isinstance(val, float) and pd.isna(val)):
        cell.font = FLAT_FONT
    elif val > 0:
        cell.font = POS_FONT
    elif val < 0:
        cell.font = NEG_FONT
    else:
        cell.font = FLAT_FONT


def _set_header(cell, text: str):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _apply_row_fill(ws, row: int, fill: PatternFill, n_cols: int = 12):
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


# ──────────────────────────────────────────────────────────
# 섹터 영역 작성 (rows 2-8)
# ──────────────────────────────────────────────────────────

def _write_sector_area(
    ws,
    sector_df: pd.DataFrame,
    advances: int,
    declines: int,
    data_date: str,
    prev_date: str,
):
    sec_map = dict(zip(sector_df["sector"], sector_df["return_1d"]))

    # 날짜 행
    ws.cell(1, 1).value = data_date
    ws.cell(1, 1).font  = DATE_FONT
    ws.cell(1, 7).value = prev_date
    ws.cell(1, 7).font  = DATE_FONT
    ws.row_dimensions[1].height = 16

    # 좌측 섹터 (A-B열, rows 2-6)
    for i, sec_name in enumerate(SECTOR_LEFT_COLS):
        r = i + 2
        ws.cell(r, 1).value     = sec_name
        ws.cell(r, 1).font      = Font(size=9)
        ws.cell(r, 1).fill      = SECT_BG
        ws.cell(r, 1).alignment = LEFT
        val = sec_map.get(sec_name)
        _set_pct_cell(ws.cell(r, 2), val)
        ws.cell(r, 2).font = POS_FONT if (val or 0) > 0 else (
            NEG_FONT if (val or 0) < 0 else FLAT_FONT
        )

    # 우측 섹터 (D-E열, rows 2-7)
    for i, sec_name in enumerate(SECTOR_RIGHT_COLS):
        r = i + 2
        ws.cell(r, 4).value     = sec_name
        ws.cell(r, 4).font      = Font(size=9)
        ws.cell(r, 4).fill      = SECT_BG
        ws.cell(r, 4).alignment = LEFT
        val = sec_map.get(sec_name)
        _set_pct_cell(ws.cell(r, 5), val)
        ws.cell(r, 5).font = POS_FONT if (val or 0) > 0 else (
            NEG_FONT if (val or 0) < 0 else FLAT_FONT
        )

    # SPX 타이틀 (G-J열, rows 2-6)
    ws.merge_cells("G2:J5")
    spx_cell = ws["G2"]
    spx_cell.value     = "SPX"
    spx_cell.font      = TITLE_FONT
    spx_cell.fill      = SPX_FILL
    spx_cell.alignment = CENTER

    # 상승/하락 (G-J열, rows 7-8)
    ws.merge_cells("G6:H6")
    ws["G6"].value     = "상승"
    ws["G6"].font      = ADV_FONT
    ws["G6"].alignment = CENTER
    ws["I6"].value     = advances
    ws["I6"].font      = ADV_FONT
    ws["I6"].alignment = CENTER

    ws.merge_cells("G7:H7")
    ws["G7"].value     = "하락"
    ws["G7"].font      = DEC_FONT
    ws["G7"].alignment = CENTER
    ws["I7"].value     = declines
    ws["I7"].font      = DEC_FONT
    ws["I7"].alignment = CENTER


# ──────────────────────────────────────────────────────────
# 테이블 헤더 작성 (rows 9-10)
# ──────────────────────────────────────────────────────────

def _write_table_headers(ws, header_row: int):
    ws.merge_cells(f"E{header_row}:H{header_row}")
    _set_header(ws.cell(header_row, 1), "")
    _set_header(ws.cell(header_row, 2), "종목코드")
    _set_header(ws.cell(header_row, 3), "종목명")
    ws.merge_cells(f"D{header_row}:D{header_row+1}")
    _set_header(ws.cell(header_row, 4), "시가총액\n(십억달러)")
    ws.cell(header_row, 4).alignment = CENTER
    _set_header(ws.cell(header_row, 5), "수익률")
    _set_header(ws.cell(header_row, 9), "시총\n순위")
    _set_header(ws.cell(header_row, 10), "섹터")
    _set_header(ws.cell(header_row, 11), "주요 사업 요약")
    _set_header(ws.cell(header_row, 12), "등락 이유 (뉴스 근거)")

    sub = header_row + 1
    _set_header(ws.cell(header_row, 2), "종목코드")
    ws.merge_cells(f"A{header_row}:A{sub}")
    ws.merge_cells(f"B{header_row}:B{sub}")
    ws.merge_cells(f"C{header_row}:C{sub}")
    ws.merge_cells(f"I{header_row}:I{sub}")
    ws.merge_cells(f"J{header_row}:J{sub}")
    ws.merge_cells(f"K{header_row}:K{sub}")
    ws.merge_cells(f"L{header_row}:L{sub}")

    for col, label in zip([5, 6, 7, 8], ["1일", "1주", "1개월", "3개월"]):
        _set_header(ws.cell(sub, col), label)


# ──────────────────────────────────────────────────────────
# 종목 데이터 행 작성
# ──────────────────────────────────────────────────────────

def _write_stock_row(ws, row: int, rec: dict, fill: PatternFill):
    _apply_row_fill(ws, row, fill)

    cells = [
        (1,  rec.get("day_rank", ""),   CENTER, Font(size=9, color="666666")),
        (2,  rec.get("ticker", ""),     CENTER, Font(size=9, bold=True)),
        (3,  rec.get("name", ""),       LEFT,   Font(size=9)),
        (4,  rec.get("market_cap_b"),   RIGHT,  Font(size=9)),
        (9,  rec.get("mc_rank", ""),    CENTER, Font(size=9, color="555555")),
        (10, rec.get("sector", ""),     LEFT,   Font(size=8, color="333333")),
        (11, rec.get("business_summary", ""), LEFT, Font(size=8, color="333333")),
        (12, rec.get("move_reason", ""), LEFT, Font(size=8, color="333333")),
    ]
    for col, val, align, font in cells:
        c = ws.cell(row, col)
        c.value     = val if val is not None else "—"
        c.alignment = align
        c.font      = font
        c.border    = THIN_BORDER

    for col, key in [(5, "return_1d"), (6, "return_1w"),
                     (7, "return_1m"), (8, "return_3m")]:
        _set_pct_cell(ws.cell(row, col), rec.get(key))

    ws.cell(row, 11).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row, 12).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 42


# ──────────────────────────────────────────────────────────
# 구분선 행
# ──────────────────────────────────────────────────────────

def _write_separator(ws, row: int):
    for col in range(1, 13):
        c = ws.cell(row, col)
        c.fill   = PatternFill("solid", fgColor="CCCCCC")
        c.border = Border(
            top    = Side(style="medium", color="888888"),
            bottom = Side(style="medium", color="888888"),
        )
    ws.row_dimensions[row].height = 4


def _write_market_summary(ws, row: int, market_summary: dict):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    title = ws.cell(row, 1)
    title.value = "📈 금일 시황 요약 — " + market_summary["headline"]
    title.font = Font(bold=True, size=11, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = LEFT

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 3, end_column=12)
    body = ws.cell(row + 1, 1)
    body.value = (
        f"관측: {market_summary['observation']}\n"
        f"해석: {market_summary['interpretation']}\n"
        f"유의: {market_summary['disclaimer']}"
    )
    body.font = Font(size=9, color="333333")
    body.fill = PatternFill("solid", fgColor="EAF2F8")
    body.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    body.border = THIN_BORDER
    ws.row_dimensions[row + 1].height = 58


# ──────────────────────────────────────────────────────────
# 메인 함수
# ──────────────────────────────────────────────────────────

def write_excel(
    top_df: pd.DataFrame,
    bottom_df: pd.DataFrame,
    sector_df: pd.DataFrame,
    advances: int,
    declines: int,
    output_path: Path,
    data_date: str = "",
    prev_date: str = "",
    market_summary: dict | None = None,
):
    """
    Excel 파일을 생성한다. 원본 레이아웃(섹터 + SPX 타이틀 + 상/하위 표)을 재현.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SPX Daily"

    # 열 너비 설정
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── 헤더 영역 ──
    _write_sector_area(ws, sector_df, advances, declines, data_date, prev_date)

    # ── 테이블 헤더 ──
    HEADER_ROW = 9
    _write_table_headers(ws, HEADER_ROW)

    # ── 상위 종목 ──
    DATA_START = HEADER_ROW + 2
    for i, (_, row_data) in enumerate(top_df.iterrows()):
        _write_stock_row(ws, DATA_START + i, row_data.to_dict(), TOP_ROW_FILL)

    # ── 구분선 ──
    SEP_ROW = DATA_START + len(top_df)
    _write_separator(ws, SEP_ROW)

    # ── 하위 종목 ──
    BOT_START = SEP_ROW + 1
    for i, (_, row_data) in enumerate(bottom_df.iterrows()):
        _write_stock_row(ws, BOT_START + i, row_data.to_dict(), BOT_ROW_FILL)

    if market_summary:
        _write_market_summary(ws, BOT_START + len(bottom_df) + 2, market_summary)

    # ── 행 고정 (스크롤 시 헤더 유지) ──
    ws.freeze_panes = f"A{DATA_START}"

    wb.save(output_path)
    logger.info("Excel 저장 완료: %s", output_path)
